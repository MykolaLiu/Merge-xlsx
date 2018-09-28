from openpyxl import load_workbook

def nothing(*argc, **kwargs):
        pass

LOG = print
ERROR = nothing#print
DEBUG = print#nothing

class BirthDate():
        def __init__(self, year, month=None, number=None):
                self._year = year
                self._month = month
                self._number = number

        @property
        def year(self):
                return self._year

        @property
        def month(self):
                return self._month


        @property
        def number(self):
                return self._number

        def __str__(self):
                return ("{} {} {}".format(self._number, self._month, self._year))

        def __eq__(self, lhc):
                return lhc.year == self.year

class DistStrategy():
        def __init__(self):
                #TODO FIGURE OUT
                pass

        def __call__(self, cls, row):
                for i in range(5):
                        if not row[i].value:
                                raise NameError("Requierd field not present row[{}]: {}".format(i, row[i].value))
                cls.id = 0
                originalid = row[0].value
                cls.surname = row[1].value.replace("'","").replace("`","")
                cls.name  = row[2].value.replace("'","").replace("`","")
                cls.patronymic = row[3].value.replace("'","").replace("`","")
                number = row[4].value
                month = row[5].value
                year = row[6].value
                cls.birth_date = BirthDate(year, month, number)
                cls.extended_date = row[7].value
                cls.gender = row[8].value
                cls.district = row[9].value
                cls.town = row[10].value
                cls.street = row[11].value
                cls.building = row[12].value
                cls.campus = row[13].value
                cls.apartment = row[14].value
                cls.landlord = row[15].value
                cls.school = None
                try:
                        cls.id = int(int(year)*1e+6 +  int(originalid))
                except:
                        raise TypeError("Requierd field has wrong type year : {} originalid: {}".format(year, originalid))

class SchoolStrategy():
        def __init__(self):
                self._counter = 0
                pass

        def inc(self):
                self._counter += 1
                return self._counter

        def __call__(self, cls, row):
                for i in range(2):
                        if not row[i].value:
                                raise NameError("Requierd field not present row[{}]: {}".format(i, row[i].value))
                cls.id = 0
                originalid = self.inc()
                fnp = row[0].value
                fnp_list = [t.strip() for t in fnp.split(" ")]
                if(len(fnp_list) != 3):
                        raise NameError("Unexpected name {} format has len {} at {}".format(fnp, len(fnp_list), originalid))
                cls.surname = fnp_list[0].replace("'","").replace("`","")
                cls.name  = fnp_list[1].replace("'","").replace("`","")
                cls.patronymic = fnp_list[2].replace("'","").replace("`","")
                cls.group = row[1].value
                cls.identity = row[2].value
                cls.gender = row[3].value
                year = row[4].value
                cls.birth_date = BirthDate(year)
                cls.gov = row[5].value
                cls.school = row[6].value
                cls.belong = row[7].value
                cls.num_district = row[7].value
                try:
                        cls.id = int(int(year)*1e+6 +  int(originalid))
                except:
                        raise TypeError("Requierd field has wrong type year : {} originalid: {}".format(year, originalid))
                cls.linked = False


class Person():
        def __init__(self, row, strategy = None):
                if strategy:
                        strategy(self, row);

                #self._id = row[0].value
                #self._name = row[1].value
                #self._surname = row[2].value
                #self._school = None if len(row) < 4 else row[3].value

        def __eq__(self, lhc):
                if ((self.name in lhc.name or lhc.name in self.name) and
                    self.surname == lhc.surname and self.birth_date == lhc.birth_date):
                        self.school = lhc.school
                        self.linked = True
                        return True
                else:
                        return False

        def __str__(self):
                formated = ""
                for key, value in self.__dict__.items():
                        formated += "	{}: {}\n".format(key, value)
                return formated

        # @property
        # def name(self):
        #         return self._name

        # @property
        # def surname(self):
        #         return self._surname

        # @property
        # def school(self):
        #            return self._school

        # @school.setter
        # def school(self, value):
        #         self._school = value if not self._school else self._school

        @property
        def dump(self):
                return [self._id, self._name, self._surname, self._school]



class Students(list):
        def __init__(self, book, strategy=None):
                self._book = book
                for row in book.rows:
                        try:
                                pr = Person(row, strategy)
                                self.append(pr)
                        except (NameError, TypeError) as err:
                                ERROR("Exception is handled {}".format(err))
                                continue

        def sheet(self, wb, ws='dummy'):
                internal_sheet = wb.create_sheet(ws, 0)
                for row in self:
                        internal_sheet.append(row.dump)
                return internal_sheet

def process(wbs, strategy, path):
        wbs = [load_workbook(path + i) for i in wbs]
        res = []
        for wb in wbs:
                sheets = wb.sheetnames
                for sh in sheets:
                        res.append(Students(wb[sh], strategy))
        return res


def entry_point():
        wb = load_workbook('./Exel/Syxiv_2000.xlsx')
        path = "./Exel/"
        if(True):
                wbs = ["Syxiv_2000.xlsx", "Syxiv_2001.xlsx", "Syxiv_2002.xlsx", "Syxiv_2003.xlsx", "Syxiv_2004.xlsx", "Syxiv_2005.xlsx", "Syxiv_2006.xlsx", "Syxiv_2007.xlsx", "Syxiv_2008.xlsx", "Syxiv_2009.xlsx", "Syxiv_2010.xlsx", "Syxiv_2011.xlsx", "Syxiv_2012.xlsx", "Syxiv_2013.xlsx"]
                #wbs = wbs[:1]
                strategy = DistStrategy()
                town_students = process(wbs, strategy, path)
        if(True):
                wbs = ['town.xlsx']
                strategy = SchoolStrategy()
                school_students = process(wbs, strategy, path)

        for students_by_school in school_students:
                for p in students_by_school:
                        for students_by_years in town_students:
                                if p in students_by_years:
                                        break
        counter = 0
        not_belong_counter = 0
        for students_by_school in school_students:
                for p in students_by_school:
                        if not p.linked:
                                counter += 1
                        if not p.linked and p.belong == 'так':
                                not_belong_counter += 1
        LOG("TOTAL NOT FOUND {}".format(counter))
        LOG("TOTAL NOT WITHOUT RESIDENT FOUND {}".format(not_belong_counter))



if __name__ == '__main__':
        entry_point()
